"""Excel file extractor."""

import asyncio
import logging
import tempfile
import time
from pathlib import Path
from typing import Any

from evidence_repository.extraction.base import BaseExtractor, ExtractionArtifact, TableData

logger = logging.getLogger(__name__)


class ExcelExtractor(BaseExtractor):
    """Extractor for Excel files (.xlsx, .xls).

    Parses Excel workbooks into structured tables with:
    - Multiple sheet support
    - Cell value type preservation
    - Sheet metadata extraction
    """

    @property
    def name(self) -> str:
        return "excel"

    @property
    def supported_content_types(self) -> list[str]:
        return [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
            "application/vnd.ms-excel",  # .xls
        ]

    @property
    def version(self) -> str:
        return "1.0.0"

    async def extract(
        self,
        data: bytes,
        filename: str,
        content_type: str,
        output_dir: Path | None = None,
    ) -> ExtractionArtifact:
        """Extract content from an Excel file.

        Args:
            data: Raw Excel bytes.
            filename: Original filename.
            content_type: MIME type.
            output_dir: Not used for Excel extraction.

        Returns:
            ExtractionArtifact with table data for each sheet.
        """
        start_time = time.time()
        artifact = self._create_artifact()

        # Parse workbook
        tables, metadata, warnings = await self._parse_workbook(data, filename)

        # Generate text representation
        text_content = self._tables_to_text(tables)

        # Calculate processing time
        processing_time_ms = int((time.time() - start_time) * 1000)

        # Build artifact
        artifact.text = text_content
        artifact.tables = tables
        artifact.metadata = metadata
        artifact.warnings = warnings
        artifact.page_count = len(tables)  # Use sheet count as page count
        artifact.processing_time_ms = processing_time_ms

        return artifact

    async def _parse_workbook(
        self,
        data: bytes,
        filename: str,
    ) -> tuple[list[TableData], dict[str, Any], list[str]]:
        """Parse Excel workbook into tables.

        Args:
            data: Raw Excel bytes.
            filename: Original filename.

        Returns:
            Tuple of (tables, metadata, warnings).
        """
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed, cannot extract Excel files")
            return [], {"error": "openpyxl not installed"}, ["openpyxl not installed"]

        def parse() -> tuple[list[TableData], dict[str, Any], list[str]]:
            warnings: list[str] = []
            tables: list[TableData] = []

            # Write to temp file (openpyxl needs a file)
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                f.write(data)
                temp_path = Path(f.name)

            try:
                workbook = openpyxl.load_workbook(temp_path, data_only=True, read_only=True)

                # Extract workbook metadata
                metadata: dict[str, Any] = {
                    "filename": filename,
                    "sheet_names": workbook.sheetnames,
                    "sheet_count": len(workbook.sheetnames),
                }

                # Extract each sheet
                for table_index, sheet_name in enumerate(workbook.sheetnames):
                    try:
                        sheet = workbook[sheet_name]
                        table = self._parse_sheet(sheet, sheet_name, table_index)
                        if table.rows:  # Only include non-empty sheets
                            tables.append(table)
                    except Exception as e:
                        warnings.append(f"Error parsing sheet '{sheet_name}': {e}")
                        logger.warning(f"Error parsing sheet '{sheet_name}': {e}")

                workbook.close()

            finally:
                temp_path.unlink(missing_ok=True)

            return tables, metadata, warnings

        return await asyncio.to_thread(parse)

    def _parse_sheet(
        self,
        sheet: Any,
        sheet_name: str,
        table_index: int,
    ) -> TableData:
        """Parse a single Excel sheet into TableData.

        Args:
            sheet: openpyxl worksheet object.
            sheet_name: Name of the sheet.
            table_index: Index of the sheet in the workbook.

        Returns:
            TableData representing the sheet.
        """
        rows_data: list[list[Any]] = []
        headers: list[str] = []
        max_col = 0

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            # Track max columns
            row_list = list(row)
            if len(row_list) > max_col:
                max_col = len(row_list)

            if row_idx == 0:
                # First non-empty row is headers
                headers = [str(cell) if cell is not None else f"Column_{i + 1}" for i, cell in enumerate(row_list)]
            else:
                # Data row
                normalized_row = [self._normalize_cell(cell) for cell in row_list]
                rows_data.append(normalized_row)

        # Ensure all rows have same length
        for row in rows_data:
            while len(row) < len(headers):
                row.append(None)

        return TableData(
            headers=headers,
            rows=rows_data,
            sheet_name=sheet_name,
            table_index=table_index,
            metadata={
                "row_count": len(rows_data),
                "column_count": len(headers),
            },
        )

    def _normalize_cell(self, cell: Any) -> Any:
        """Normalize a cell value.

        Args:
            cell: Raw cell value from openpyxl.

        Returns:
            Normalized value.
        """
        if cell is None:
            return None

        # Handle datetime objects
        from datetime import datetime, date
        if isinstance(cell, datetime):
            return cell.isoformat()
        if isinstance(cell, date):
            return cell.isoformat()

        return cell

    def _tables_to_text(self, tables: list[TableData]) -> str:
        """Convert all tables to text representation.

        Args:
            tables: List of TableData objects.

        Returns:
            Text representation of all tables.
        """
        parts: list[str] = []

        for table in tables:
            if table.sheet_name:
                parts.append(f"=== Sheet: {table.sheet_name} ===\n")

            # Headers
            parts.append(" | ".join(table.headers))
            parts.append("-" * max(40, len(parts[-1])))

            # Rows
            for row in table.rows:
                row_strs = [str(cell) if cell is not None else "" for cell in row]
                parts.append(" | ".join(row_strs))

            parts.append("")  # Empty line between sheets

        return "\n".join(parts)
